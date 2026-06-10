import os
import zipfile
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinterdnd2 import DND_FILES
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import threading
import chardet
import re
import shutil

# -------------------------- 辅助工具函数 --------------------------
def read_txt_file(file_path):
    """读取TXT文件：自动检测编码、过滤乱码、按行去重，返回清理后的内容"""
    if not os.path.exists(file_path):
        return ""
    try:
        # 二进制读取+编码自动检测，解决中文乱码
        with open(file_path, 'rb') as f:
            raw_data = f.read()
        detect_result = chardet.detect(raw_data)
        encoding = detect_result['encoding'] if detect_result['encoding'] else 'utf-8'
        # 解码+乱码过滤：仅保留中文、英文、数字、常见标点
        content = raw_data.decode(encoding, errors='ignore')
        content = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9，。！？、；：""''（）【】\s]', '', content)
        # 按行去重，保留原始顺序
        lines = [line.strip() for line in content.splitlines() if line.strip()]
        unique_lines = list(dict.fromkeys(lines))
        return '\n'.join(unique_lines)
    except Exception as e:
        return f"读取失败：{str(e)}"

# -------------------------- 核心功能类 --------------------------

class ZipExcelTab:
    def __init__(self, notebook):
        self.tab = ttk.Frame(notebook, style="Custom.TFrame")
        notebook.add(self.tab, text="ZIP转Excel")
        self.zip_path = tk.StringVar()
        self.create_ui()

    def create_ui(self):
        # 主容器
        main_frame = ttk.Frame(self.tab, style="Custom.TFrame")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # 标题区域
        title_frame = ttk.Frame(main_frame, style="Custom.TFrame")
        title_frame.pack(fill="x", pady=(0, 20))
        
        ttk.Label(title_frame, text="ZIP文件夹转Excel", 
                 font=("微软雅黑", 18, "bold"), foreground="#333333").pack()
        ttk.Label(title_frame, text="自动解压ZIP，读取子文件夹内容并生成Excel", 
                 foreground="#666666", font=("微软雅黑", 9)).pack(pady=(5, 0))

        # 文件选择行
        file_row = ttk.Frame(main_frame, style="Custom.TFrame")
        file_row.pack(fill="x", pady=(0, 15))
        
        ttk.Label(file_row, text="ZIP 文件：", font=("微软雅黑", 10)).pack(side="left", padx=(0, 10))
        entry = tk.Entry(file_row, textvariable=self.zip_path, width=50, 
                        font=("微软雅黑", 10), relief="solid", bd=1)
        entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Button(file_row, text="选择文件", command=self.select_zip, 
                  style="Accent.TButton").pack(side="left")

        # 拖拽区域
        drop_frame = ttk.LabelFrame(main_frame, text="拖拽区域", padding="20")
        drop_frame.pack(fill="x", pady=(0, 15))
        lbl = ttk.Label(drop_frame, text="📁 拖入 ZIP 文件即可开始", 
                       font=("微软雅黑", 12), foreground="#409eff", 
                       background="#f0f7ff", padding=20)
        lbl.pack(expand=True, fill="both")
        drop_frame.drop_target_register(DND_FILES)
        drop_frame.dnd_bind('<<Drop>>', self.on_drop)

        # 进度条区域
        progress_frame = ttk.Frame(main_frame, style="Custom.TFrame")
        progress_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(progress_frame, text="进度：", font=("微软雅黑", 10)).pack(side="left", padx=(0, 10))
        self.progress = ttk.Progressbar(progress_frame, length=500, style="TProgressbar")
        self.progress.pack(side="left", fill="x", expand=True)

        # ========== 两个按钮并排放置 ==========
        btn_frame = ttk.Frame(main_frame, style="Custom.TFrame")
        btn_frame.pack(pady=15)
        self.btn = ttk.Button(btn_frame, text="✅ 开始转换", command=self.start_convert, 
                             width=20, style="Accent.TButton")
        self.btn.pack(side="left", padx=10, ipadx=10, ipady=5)
        self.btn_ext = ttk.Button(btn_frame, text="⭐ 开始转换（拓展版）", command=self.start_convert_extended,
                                 width=20, style="Accent.TButton")
        self.btn_ext.pack(side="left", padx=10, ipadx=10, ipady=5)

        # 日志区域
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="10")
        log_frame.pack(fill="both", expand=True, pady=(0, 0))
        
        self.log_text = tk.Text(log_frame, width=88, height=10, font=("微软雅黑", 9),
                               relief="solid", bd=1, bg="white")
        self.log_text.pack(fill="both", expand=True)
        self.log_text.config(state=tk.DISABLED)

        # 初始化日志
        self.log("自动解压ZIP，读取子文件夹内容并生成Excel")
        self.log("准备就绪，请选择或拖拽 ZIP 文件")
        self.log("标准版：支持5张图片，A~M列表头")
        self.log("拓展版：支持12张图片，A~Y列表头，新增音乐链接、其他文案等列")

    # 日志输出
    def log(self, msg):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    # 选择ZIP文件
    def select_zip(self):
        p = filedialog.askopenfilename(filetypes=[("ZIP 文件", "*.zip")])
        if p: 
            self.zip_path.set(p)

    # 拖拽文件处理
    def on_drop(self, e):
        p = e.data.strip("{}").replace("\\", "/")
        if p.lower().endswith(".zip"):
            self.zip_path.set(p)

    # 开始转换（标准版）
    def start_convert(self):
        zpath = self.zip_path.get()
        if not os.path.exists(zpath):
            messagebox.showerror("错误", "请选择有效的 ZIP 文件")
            return
        self.btn.config(state=tk.DISABLED)
        self.btn_ext.config(state=tk.DISABLED)
        self.progress["value"] = 0
        threading.Thread(target=self.do_convert, daemon=True).start()

    # 开始转换（拓展版）
    def start_convert_extended(self):
        zpath = self.zip_path.get()
        if not os.path.exists(zpath):
            messagebox.showerror("错误", "请选择有效的 ZIP 文件")
            return
        self.btn.config(state=tk.DISABLED)
        self.btn_ext.config(state=tk.DISABLED)
        self.progress["value"] = 0
        threading.Thread(target=self.do_convert_extended, daemon=True).start()

    # -------------------- 标准版转换逻辑（与原版一致） --------------------
    def do_convert(self):
        temp = f"temp_unzip_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        try:
            os.makedirs(temp, exist_ok=True)
            self.log("开始解压ZIP文件...")
            with zipfile.ZipFile(self.zip_path.get(), 'r') as zf:
                total_files = len(zf.infolist())
                for idx, info in enumerate(zf.infolist(), 1):
                    try:
                        filename = info.filename.encode('cp437').decode('gbk')
                    except:
                        filename = info.filename
                    info.filename = filename
                    zf.extract(info, temp)
                    self.progress["value"] = idx / total_files * 50
                    self.tab.update_idletasks()
            self.log("✅ 解压完成，开始读取文件夹内容...")

            folders = [f for f in os.listdir(temp) if os.path.isdir(os.path.join(temp, f))]
            if not folders:
                self.log("⚠️ 未找到任何子文件夹，转换终止")
                messagebox.showwarning("警告", "ZIP文件内未找到任何子文件夹")
                return
            self.log(f"找到 {len(folders)} 个有效子文件夹，开始生成Excel...")

            wb = Workbook()
            ws = wb.active
            ws.title = "ZIP转换结果"
            headers = ["序号", "标题", "文案", "话题", "大字报", "图片1", "图片2", "图片3", "图片4", "图片5", "视频", "使用说明", "指定用户"]
            ws.append(headers)

            for col in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=col)
                cell.font = cell.font.copy(bold=True)
                cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
                if col in [6,7,8,9,10]:
                    ws.column_dimensions[chr(64+col)].width = 15
                elif col in [3,4,5]:
                    ws.column_dimensions[chr(64+col)].width = 20
                else:
                    ws.column_dimensions[chr(64+col)].width = 12

            for i, folder_name in enumerate(folders, 2):
                current_idx = i - 1
                folder_path = os.path.join(temp, folder_name)
                self.log(f"正在处理第 {current_idx}/{len(folders)} 个文件夹：{folder_name}")

                ws.cell(row=i, column=1, value=current_idx)
                ws.cell(row=i, column=2, value=folder_name)
                ws.row_dimensions[i].height = 100

                # 文案列
                content_list = []
                for txt_name in ["文案.txt", "正文.txt"]:
                    txt_path = os.path.join(folder_path, txt_name)
                    if os.path.exists(txt_path):
                        content_list.append(read_txt_file(txt_path))
                if not content_list:
                    for file in os.listdir(folder_path):
                        if file.lower().endswith(".txt"):
                            txt_path = os.path.join(folder_path, file)
                            content_list.append(read_txt_file(txt_path))
                all_content = '\n'.join([c for c in content_list if c])
                unique_content = '\n'.join(list(dict.fromkeys(all_content.splitlines())))
                ws.cell(row=i, column=3, value=unique_content).alignment = \
                    ws.cell(row=i, column=3).alignment.copy(wrap_text=True, vertical="top")

                col_map = {4: "话题.txt", 5: "大字报.txt", 11: "视频.txt", 12: "使用说明.txt", 13: "指定用户.txt"}
                for col, txt_name in col_map.items():
                    content = read_txt_file(os.path.join(folder_path, txt_name))
                    ws.cell(row=i, column=col, value=content).alignment = \
                        ws.cell(row=i, column=col).alignment.copy(wrap_text=True, vertical="top")

                img_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.gif')
                img_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith(img_extensions)]
                for img_idx, img_path in enumerate(img_files[:5]):
                    col = 6 + img_idx
                    cell_coord = f"{chr(64 + col)}{i}"
                    try:
                        img_obj = Image(img_path)
                        img_obj.anchor_type = "oneCell"
                        img_obj.anchor = cell_coord
                        original_w, original_h = img_obj.width, img_obj.height
                        target_w = 15 * 96 / 10
                        target_h = 100 * 96 / 72
                        scale = min(target_w / original_w, target_h / original_h)
                        img_obj.width, img_obj.height = original_w * scale, original_h * scale
                        ws.add_image(img_obj)
                    except Exception as e:
                        self.log(f"⚠️ 第{current_idx}个文件夹的第{img_idx+1}张图片加载失败：{str(e)}")
                        continue

                self.progress["value"] = 50 + (current_idx / len(folders) * 50)
                self.tab.update_idletasks()

            output_dir = os.path.dirname(self.zip_path.get())
            output_name = f"ZIP转换结果_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            output_path = os.path.join(output_dir, output_name)
            wb.save(output_path)
            self.log(f"✅ 转换完成！文件已保存至：{output_path}")
            messagebox.showinfo("转换完成", f"Excel文件已成功生成！\n保存路径：{output_path}")
        except Exception as e:
            self.log(f"❌ 转换过程中发生错误：{str(e)}")
            messagebox.showerror("错误", f"转换失败：{str(e)}")
        finally:
            try:
                if os.path.exists(temp):
                    shutil.rmtree(temp)
                    self.log("🗑️  临时文件已清理完成")
            except Exception as e:
                self.log(f"⚠️  临时文件清理失败：{str(e)}")
            self.btn.config(state=tk.NORMAL)
            self.btn_ext.config(state=tk.NORMAL)
            self.progress["value"] = 0

    # -------------------- 拓展版转换逻辑 --------------------
    def do_convert_extended(self):
        temp = f"temp_unzip_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        try:
            os.makedirs(temp, exist_ok=True)
            self.log("开始解压ZIP文件...")
            with zipfile.ZipFile(self.zip_path.get(), 'r') as zf:
                total_files = len(zf.infolist())
                for idx, info in enumerate(zf.infolist(), 1):
                    try:
                        filename = info.filename.encode('cp437').decode('gbk')
                    except:
                        filename = info.filename
                    info.filename = filename
                    zf.extract(info, temp)
                    self.progress["value"] = idx / total_files * 50
                    self.tab.update_idletasks()
            self.log("✅ 解压完成，开始读取文件夹内容...")

            folders = [f for f in os.listdir(temp) if os.path.isdir(os.path.join(temp, f))]
            if not folders:
                self.log("⚠️ 未找到任何子文件夹，转换终止")
                messagebox.showwarning("警告", "ZIP文件内未找到任何子文件夹")
                return
            self.log(f"找到 {len(folders)} 个有效子文件夹，开始生成Excel（拓展版）...")

            wb = Workbook()
            ws = wb.active
            ws.title = "ZIP转换结果_拓展版"
            # 拓展版表头：A1~Y1 共25列
            headers = [
                "序号", "标题", "文案", "话题", "大字报",
                "图片1", "图片2", "图片3", "图片4", "图片5",
                "图片6", "图片7", "图片8", "图片9", "图片10",
                "图片11", "图片12",
                "视频", "指定音乐链接", "其他文案",
                "附件1", "附件2", "附件3", "使用说明", "指定用户"
            ]
            ws.append(headers)

            # 设置列宽和表头样式
            for col in range(1, len(headers)+1):
                cell = ws.cell(row=1, column=col)
                cell.font = cell.font.copy(bold=True)
                cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
                # 图片列（第6~17列）宽度15
                if 6 <= col <= 17:
                    ws.column_dimensions[chr(64+col)].width = 15
                # 长文本列：文案(3)、话题(4)、大字报(5)、指定音乐链接(19)、其他文案(20)、使用说明(24)、指定用户(25)
                elif col in [3,4,5,19,20,24,25]:
                    ws.column_dimensions[chr(64+col)].width = 20
                else:
                    ws.column_dimensions[chr(64+col)].width = 12

            # 遍历每个子文件夹
            for i, folder_name in enumerate(folders, 2):
                current_idx = i - 1
                folder_path = os.path.join(temp, folder_name)
                self.log(f"正在处理第 {current_idx}/{len(folders)} 个文件夹：{folder_name}")

                # 序号、标题
                ws.cell(row=i, column=1, value=current_idx)
                ws.cell(row=i, column=2, value=folder_name)
                ws.row_dimensions[i].height = 100  # 预留图片行高

                # ---------- 文案列（C列）----------
                content_list = []
                for txt_name in ["文案.txt", "正文.txt"]:
                    txt_path = os.path.join(folder_path, txt_name)
                    if os.path.exists(txt_path):
                        content_list.append(read_txt_file(txt_path))
                if not content_list:
                    for file in os.listdir(folder_path):
                        if file.lower().endswith(".txt"):
                            txt_path = os.path.join(folder_path, file)
                            content_list.append(read_txt_file(txt_path))
                all_content = '\n'.join([c for c in content_list if c])
                unique_content = '\n'.join(list(dict.fromkeys(all_content.splitlines())))
                ws.cell(row=i, column=3, value=unique_content).alignment = \
                    ws.cell(row=i, column=3).alignment.copy(wrap_text=True, vertical="top")

                # ---------- 固定列映射（话题、大字报、使用说明、指定用户）----------
                # 注意列号变化：话题4，大字报5，使用说明24，指定用户25
                col_map = {
                    4: "话题.txt",
                    5: "大字报.txt",
                    24: "使用说明.txt",
                    25: "指定用户.txt"
                }
                for col, txt_name in col_map.items():
                    content = read_txt_file(os.path.join(folder_path, txt_name))
                    ws.cell(row=i, column=col, value=content).alignment = \
                        ws.cell(row=i, column=col).alignment.copy(wrap_text=True, vertical="top")

                # ---------- 新增列：指定音乐链接（19列）、其他文案（20列）----------
                music_content = read_txt_file(os.path.join(folder_path, "指定音乐链接.txt"))
                if not music_content:
                    music_content = read_txt_file(os.path.join(folder_path, "音乐链接.txt"))
                ws.cell(row=i, column=19, value=music_content).alignment = \
                    ws.cell(row=i, column=19).alignment.copy(wrap_text=True, vertical="top")

                other_content = read_txt_file(os.path.join(folder_path, "其他文案.txt"))
                ws.cell(row=i, column=20, value=other_content).alignment = \
                    ws.cell(row=i, column=20).alignment.copy(wrap_text=True, vertical="top")

                # ---------- 视频列（18列）、附件1~3（21~23列）留空，不读取任何内容 ----------
                # （已经自动为空，无需额外操作）

                # ---------- 图片处理：最多12张，嵌入F~Q列（第6~17列） ----------
                img_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.gif')
                img_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith(img_extensions)]
                for img_idx, img_path in enumerate(img_files[:12]):  # 最多12张
                    col = 6 + img_idx  # 图片1=6, 图片12=17
                    cell_coord = f"{chr(64 + col)}{i}"
                    try:
                        img_obj = Image(img_path)
                        img_obj.anchor_type = "oneCell"
                        img_obj.anchor = cell_coord
                        original_w, original_h = img_obj.width, img_obj.height
                        target_w = 15 * 96 / 10   # 列宽15对应像素
                        target_h = 100 * 96 / 72  # 行高100对应像素
                        scale = min(target_w / original_w, target_h / original_h)
                        img_obj.width, img_obj.height = original_w * scale, original_h * scale
                        ws.add_image(img_obj)
                    except Exception as e:
                        self.log(f"⚠️ 第{current_idx}个文件夹的第{img_idx+1}张图片加载失败：{str(e)}")
                        continue

                # 更新进度
                self.progress["value"] = 50 + (current_idx / len(folders) * 50)
                self.tab.update_idletasks()

            # 保存Excel
            output_dir = os.path.dirname(self.zip_path.get())
            output_name = f"ZIP转换结果_拓展版_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            output_path = os.path.join(output_dir, output_name)
            wb.save(output_path)
            self.log(f"✅ 拓展版转换完成！文件已保存至：{output_path}")
            messagebox.showinfo("转换完成", f"拓展版Excel文件已成功生成！\n保存路径：{output_path}")
        except Exception as e:
            self.log(f"❌ 拓展版转换过程中发生错误：{str(e)}")
            messagebox.showerror("错误", f"转换失败：{str(e)}")
        finally:
            try:
                if os.path.exists(temp):
                    shutil.rmtree(temp)
                    self.log("🗑️  临时文件已清理完成")
            except Exception as e:
                self.log(f"⚠️  临时文件清理失败：{str(e)}")
            self.btn.config(state=tk.NORMAL)
            self.btn_ext.config(state=tk.NORMAL)
            self.progress["value"] = 0