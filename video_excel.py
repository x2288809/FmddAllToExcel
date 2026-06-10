import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinterdnd2 import DND_FILES
from openpyxl import Workbook

class VideoExcelTab:
    def __init__(self, notebook):
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="视频转Excel")
        self.video_folder_path = tk.StringVar(value="未选择文件夹")
        self.create_ui()

    def create_ui(self):
        # 标题
        ttk.Label(self.tab, text="视频文件名导入Excel", font=("微软雅黑", 16, "bold")).pack(pady=18)
        ttk.Label(self.tab, text="统一填写文案内容，无需可留空，生成后可在表格内修改", 
                  foreground="#555").pack(pady=2)

        # 输入区域
        input_frame = ttk.Frame(self.tab, padding="10 10 10 10")
        input_frame.pack(fill="x", padx=35, pady=8)

        labels = ["标题", "正文", "话题", "大字报", "使用说明", "指定账号", "起始序号"]
        self.entries = {}

        for idx, text in enumerate(labels):
            ttk.Label(input_frame, text=f"{text}：", width=12, font=("微软雅黑", 10)).grid(
                row=idx, column=0, sticky="w", pady=6)
            entry = ttk.Entry(input_frame, font=("微软雅黑", 10))
            entry.grid(row=idx, column=1, sticky="ew", padx=4)
            self.entries[text] = entry

        self.entries["起始序号"].insert(0, "1")
        input_frame.columnconfigure(1, weight=1)

        # 文件夹选择
        ttk.Button(self.tab, text="选择视频文件夹", command=self.select_folder, 
                  style="Accent.TButton").pack(pady=6)
        ttk.Label(self.tab, textvariable=self.video_folder_path, font=("微软雅黑", 9)).pack()

        # 拖拽区域
        drop_frame = ttk.LabelFrame(self.tab, text="拖拽文件夹到此处", padding="10")
        drop_frame.pack(padx=35, pady=10, fill="x")
        drop_label = ttk.Label(drop_frame, text="点击选择 或 直接拖拽文件夹", font=("微软雅黑", 10))
        drop_label.pack(expand=True)
        drop_label.drop_target_register(DND_FILES)
        drop_label.dnd_bind('<<Drop>>', self.on_drop)

        # 生成按钮
        self.btn_start = ttk.Button(self.tab, text="✅ 开始生成Excel", command=self.start_generate, 
                                   state=tk.DISABLED, style="Accent.TButton")
        self.btn_start.pack(pady=25, ipadx=30, ipady=6)

    def select_folder(self):
        path = filedialog.askdirectory(title="选择视频文件夹")
        if path:
            self.video_folder_path.set(path)
            self.btn_start.config(state=tk.NORMAL)

    def on_drop(self, event):
        path = event.data.strip("{}")
        if os.path.isdir(path):
            self.video_folder_path.set(path)
            self.btn_start.config(state=tk.NORMAL)

    def start_generate(self):
        folder = self.video_folder_path.get()
        if not os.path.isdir(folder):
            messagebox.showerror("错误", "请选择有效文件夹")
            return

        title = self.entries["标题"].get().strip()
        content = self.entries["正文"].get().strip()
        topic = self.entries["话题"].get().strip()
        poster = self.entries["大字报"].get().strip()
        desc = self.entries["使用说明"].get().strip()
        account = self.entries["指定账号"].get().strip()
        
        try:
            start_serial = max(int(self.entries["起始序号"].get().strip()), 1)
        except:
            start_serial = 1

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "视频列表"
            headers = ["序号", "标题", "正文", "话题", "大字报", "图片1", "图片2", 
                       "图片3", "图片4", "图片5", "视频", "使用说明", "指定账号"]
            for i, h in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=h)

            exts = (".mp4", ".avi", ".mov", ".mkv", ".flv", ".wmv", ".m4v")
            videos = [f for f in os.listdir(folder) if f.lower().endswith(exts)]

            for idx, video in enumerate(videos, 2):
                serial = start_serial + (idx - 2)
                ws.cell(row=idx, column=1, value=serial)
                ws.cell(row=idx, column=2, value=title)
                ws.cell(row=idx, column=3, value=content)
                ws.cell(row=idx, column=4, value=topic)
                ws.cell(row=idx, column=5, value=poster)
                ws.cell(row=idx, column=11, value=video)
                ws.cell(row=idx, column=12, value=desc)
                ws.cell(row=idx, column=13, value=account)

            save_path = os.path.join(folder, "视频内容表格.xlsx")
            wb.save(save_path)
            messagebox.showinfo("✅ 完成", f"成功生成 {len(videos)} 条数据")
            os.startfile(folder)
        except Exception as e:
            messagebox.showerror("❌ 失败", str(e))