import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinterdnd2 import DND_FILES
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re

class ExtendExcelTab:
    def __init__(self, notebook):
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="Excel拓展表格生成")
        # 路径变量
        self.extend_excel_path = tk.StringVar(value="未选择Excel拓展表")
        self.video_folder = tk.StringVar(value="未选择视频文件夹")
        self.attach1_folder = tk.StringVar(value="未选择附件1文件夹")
        self.attach2_folder = tk.StringVar(value="未选择附件2文件夹")
        self.attach3_folder = tk.StringVar(value="未选择附件3文件夹")
        
        self.create_ui()

    def create_ui(self):
        # 顶部标题
        ttk.Label(self.tab, text="Excel拓展表格生成", font=("微软雅黑", 16, "bold")).pack(pady=10)
        ttk.Label(self.tab, text="支持多列文件批量填充，生成拓展Excel表格；拖入已有拓展版，则会更新现有数据", foreground="#555").pack(pady=2)

        # ========== 第一部分：文本输入区域 ==========
        input_frame = ttk.Frame(self.tab, padding="10 10 10 10")
        input_frame.pack(fill="x", padx=30, pady=8)

        base_items = [
            ("标题：", "title"),
            ("文案：", "content"),
            ("话题：", "topic"),
            ("大字报：", "poster"),
            ("指定音乐链接：", "music_url"),
            ("其他文案：", "other_content"),
            ("使用说明：", "instruct"),
            ("起始序号：", "start_idx")
        ]
        self.entries = {}
        for idx, (text, key) in enumerate(base_items):
            ttk.Label(input_frame, text=text, width=14, font=("微软雅黑", 10)).grid(
                row=idx, column=0, sticky="w", pady=4)
            entry = ttk.Entry(input_frame, font=("微软雅黑", 10))
            entry.grid(row=idx, column=1, sticky="ew", padx=4)
            self.entries[key] = entry
            if key == "start_idx":
                entry.insert(0, "1")
        input_frame.columnconfigure(1, weight=1)

        # ========== 第二部分：拖拽区域（横向布局） ==========
        # 1. 第一行：Excel拓展表 + 视频 两列并排
        row1_frame = ttk.Frame(self.tab)
        row1_frame.pack(fill="x", padx=30, pady=6)
        row1_frame.columnconfigure(0, weight=1)
        row1_frame.columnconfigure(1, weight=1)

        # Excel拓展表 区域
        self.build_drop_panel(row1_frame, 0, "Excel拓展表", self.extend_excel_path)
        # 视频 区域
        self.build_drop_panel(row1_frame, 1, "视频", self.video_folder)

        # 2. 第二行：附件1 + 附件2 + 附件3 三列并排
        row2_frame = ttk.Frame(self.tab)
        row2_frame.pack(fill="x", padx=30, pady=6)
        row2_frame.columnconfigure(0, weight=1)
        row2_frame.columnconfigure(1, weight=1)
        row2_frame.columnconfigure(2, weight=1)

        self.build_drop_panel(row2_frame, 0, "附件1", self.attach1_folder)
        self.build_drop_panel(row2_frame, 1, "附件2", self.attach2_folder)
        self.build_drop_panel(row2_frame, 2, "附件3", self.attach3_folder)

        # ========== 底部生成按钮 ==========
        self.btn_generate = ttk.Button(
            self.tab, text="✅ 生成/更新拓展Excel表格",
            command=self.generate_extend_excel,
            style="Accent.TButton"
        )
        self.btn_generate.pack(pady=20, ipadx=30, ipady=6)

    def build_drop_panel(self, parent, col, label_text, var):
        """封装：构建单个拖拽面板（横向布局复用）"""
        frame = ttk.LabelFrame(parent, text=label_text, padding="8")
        frame.grid(row=0, column=col, sticky="nsew", padx=4)

        lbl_tip = ttk.Label(frame, text="拖拽/点击选择", font=("微软雅黑", 9))
        lbl_tip.pack(pady=8)

        lbl_path = ttk.Label(frame, textvariable=var, font=("微软雅黑", 8), foreground="#666")
        lbl_path.pack(fill="x", padx=4)

        # 拖拽+点击事件
        lbl_tip.drop_target_register(DND_FILES)
        lbl_tip.dnd_bind('<<Drop>>', lambda e, v=var: self.on_drop(e, v))
        lbl_tip.bind("<Button-1>", lambda e, v=var, t=label_text: self.select_file_folder(e, v, t))

    def select_file_folder(self, event, var, type_name):
        if type_name == "Excel拓展表":
            path = filedialog.askopenfilename(
                title="选择Excel拓展表",
                filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
            )
        else:
            path = filedialog.askdirectory(title=f"选择{type_name}文件夹")
        if path:
            var.set(path)

    def on_drop(self, event, var):
        path = event.data.strip("{}")
        if os.path.exists(path):
            var.set(path)
    
    def get_file_list(self, folder_path, filter_type=None):
        if not os.path.isdir(folder_path):
            return []
        
        video_exts = ('.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv', '.webm')
        file_list = []
        
        for f in os.listdir(folder_path):
            file_path = os.path.join(folder_path, f)
            if os.path.isfile(file_path):
                if filter_type == "video" and not f.lower().endswith(video_exts):
                    continue
                file_list.append(f)

        # ========== 新增：自然排序（数字正确排序） ==========
        def natural_sort_key(s):
            # 把字符串里的数字转成整数，非数字保持原样
            return [int(text) if text.isdigit() else text.lower()
                    for text in re.split(r'(\d+)', s)]
        
        file_list.sort(key=natural_sort_key)
        # ====================================================
        
        return file_list

    def generate_extend_excel(self):
        try:
            start_idx = int(self.entries["start_idx"].get().strip())
        except:
            messagebox.showerror("错误", "起始序号必须是有效数字")
            return

        title = self.entries["title"].get().strip()
        content = self.entries["content"].get().strip()
        topic = self.entries["topic"].get().strip()
        poster = self.entries["poster"].get().strip()
        music_url = self.entries["music_url"].get().strip()
        other_content = self.entries["other_content"].get().strip()
        instruct = self.entries["instruct"].get().strip()

        headers = [
            "序号", "标题", "文案", "话题", "大字报",
            "图片1", "图片2", "图片3", "图片4", "图片5",
            "图片6", "图片7", "图片8", "图片9", "图片10",
            "图片11", "图片12", "视频", "指定音乐链接", "其他文案",
            "附件1", "附件2", "附件3", "使用说明", "指定用户"
        ]

        excel_path = self.extend_excel_path.get()
        if excel_path == "未选择Excel拓展表":
            wb = Workbook()
            ws = wb.active
            ws.title = "拓展文件表格"
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            save_path = os.path.join(os.getcwd(), "拓展文件表格.xlsx")
        else:
            if not os.path.isfile(excel_path) or not excel_path.endswith(('.xlsx', '.xls')):
                messagebox.showerror("错误", "选择的Excel拓展表无效")
                return
            wb = load_workbook(excel_path)
            ws = wb.active
            save_path = excel_path

        video_files = self.get_file_list(self.video_folder.get(), filter_type="video")
        attach1_files = self.get_file_list(self.attach1_folder.get())
        attach2_files = self.get_file_list(self.attach2_folder.get())
        attach3_files = self.get_file_list(self.attach3_folder.get())

        max_rows = max(len(video_files), len(attach1_files), len(attach2_files), len(attach3_files), 1)

        for row_idx in range(2, max_rows + 2):
            current_idx = start_idx + (row_idx - 2)
            ws.cell(row=row_idx, column=1, value=current_idx)
            ws.cell(row=row_idx, column=2, value=title)
            ws.cell(row=row_idx, column=3, value=content)
            ws.cell(row=row_idx, column=4, value=topic)
            ws.cell(row=row_idx, column=5, value=poster)
            ws.cell(row=row_idx, column=19, value=music_url)
            ws.cell(row=row_idx, column=20, value=other_content)
            ws.cell(row=row_idx, column=24, value=instruct)

            if row_idx - 2 < len(video_files):
                ws.cell(row=row_idx, column=18, value=video_files[row_idx - 2])
            if row_idx - 2 < len(attach1_files):
                ws.cell(row=row_idx, column=21, value=attach1_files[row_idx - 2])
            if row_idx - 2 < len(attach2_files):
                ws.cell(row=row_idx, column=22, value=attach2_files[row_idx - 2])
            if row_idx - 2 < len(attach3_files):
                ws.cell(row=row_idx, column=23, value=attach3_files[row_idx - 2])

        col_widths = [6, 20, 30, 15, 15,
                      15, 15, 15, 15, 15,
                      15, 15, 15, 15, 15,
                      15, 15, 25, 30, 20,
                      25, 25, 25, 20, 15]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        try:
            wb.save(save_path)
            messagebox.showinfo("✅ 成功", f"拓展Excel表格生成/更新完成！\n文件路径：{save_path}")
            if os.path.exists(save_path):
                os.startfile(os.path.dirname(save_path))
        except Exception as e:
            messagebox.showerror("❌ 失败", f"保存Excel失败：{str(e)}")