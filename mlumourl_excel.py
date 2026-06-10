import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinterdnd2 import DND_FILES
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

class UrlExcelTab:
    def __init__(self, notebook):
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="MlumoURL图片转Excel")
        self.file_path = tk.StringVar(value="未选择Excel文件")
        self.create_ui()

    def create_ui(self):
        ttk.Label(self.tab, text="Mlumo 图片URL转Excel", font=("微软雅黑", 16, "bold")).pack(pady=18)
        ttk.Label(self.tab, text="自动下载URL图片并生成带预览的Excel表格", foreground="#555").pack(pady=2)

        ttk.Label(self.tab, textvariable=self.file_path, font=("微软雅黑", 10)).pack(pady=6)
        ttk.Button(self.tab, text="选择Excel文件", command=self.select_file, style="Accent.TButton").pack(pady=6)

        self.progress = ttk.Progressbar(self.tab, length=500)
        self.progress.pack(pady=10)
        self.status = ttk.Label(self.tab, text="等待操作...", font=("微软雅黑", 10))
        self.status.pack(pady=6)

        self.btn = ttk.Button(self.tab, text="✅ 开始转换", command=self.start, style="Accent.TButton")
        self.btn.pack(pady=20, ipadx=30, ipady=6)

        self.tab.drop_target_register(DND_FILES)
        self.tab.dnd_bind('<<Drop>>', self.on_drop)

    def select_file(self):
        p = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx;*.xls")])
        if p: self.file_path.set(p)

    def on_drop(self, e):
        p = e.data.strip("{}")
        if p.lower().endswith((".xlsx", ".xls")):
            self.file_path.set(p)

    def start(self):
        p = self.file_path.get()
        if not os.path.exists(p):
            messagebox.showwarning("提示", "请先选择Excel文件")
            return
        self.btn.config(state=tk.DISABLED)
        self.status.config(text="正在处理...")

        temp = "temp_img"
        os.makedirs(temp, exist_ok=True)

        try:
            wb_in = load_workbook(p)
            ws_in = wb_in.active
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.append(["序号", "标题", "正文", "图片1", "图片2", "图片3", "图片4", "图片5"])

            total = max(ws_in.max_row - 1, 1)
            for row in range(2, ws_in.max_row + 1):
                title = ws_in[f"B{row}"].value
                url = ws_in[f"F{row}"].value
                ws_out.cell(row=row, column=1, value=row-1)
                ws_out.cell(row=row, column=2, value=title)

                if url and str(url).startswith("http"):
                    try:
                        r = requests.get(url, timeout=15)
                        img_path = os.path.join(temp, f"{row}.jpg")
                        with open(img_path, 'wb') as f:
                            f.write(r.content)
                        ws_out.add_image(Image(img_path), f"F{row}")
                    except Exception as e:
                        pass

                self.progress["value"] = (row-1)/total*100
                self.tab.update_idletasks()

            out = os.path.join(os.path.dirname(p), "URL图片转换结果.xlsx")
            wb_out.save(out)
            self.status.config(text="✅ 处理完成")
            messagebox.showinfo("成功", f"文件已保存：\n{out}")
        except Exception as e:
            messagebox.showerror("失败", str(e))
        finally:
            self.btn.config(state=tk.NORMAL)