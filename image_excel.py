import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinterdnd2 import DND_FILES
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

class ImageExcelTab:
    def __init__(self, notebook):
        self.tab = ttk.Frame(notebook)
        notebook.add(self.tab, text="图片批量转Excel")
        self.folder_path = tk.StringVar(value="未选择文件夹")
        self.create_ui()

    def create_ui(self):
        ttk.Label(self.tab, text="图片批量转Excel", font=("微软雅黑", 16, "bold")).pack(pady=18)
        ttk.Label(self.tab, text="自动读取文件夹图片，生成带预览的Excel表格", 
                  foreground="#555").pack(pady=2)

        input_frame = ttk.Frame(self.tab, padding="10 10 10 10")
        input_frame.pack(fill="x", padx=35, pady=8)

        items = [("标题", "title"), ("文案", "content"), ("话题", "topic"), 
                 ("大字报", "poster"), ("使用说明", "instruct"), ("每行图片数(1-5)", "per_row")]
        self.entries = {}

        for idx, (text, key) in enumerate(items):
            ttk.Label(input_frame, text=f"{text}：", width=16, font=("微软雅黑", 10)).grid(
                row=idx, column=0, sticky="w", pady=6)
            entry = ttk.Entry(input_frame, font=("微软雅黑", 10))
            entry.grid(row=idx, column=1, sticky="ew", padx=4)
            self.entries[key] = entry

        self.entries["per_row"].insert(0, "1")
        input_frame.columnconfigure(1, weight=1)

        ttk.Button(self.tab, text="选择图片文件夹", command=self.select_folder, 
                  style="Accent.TButton").pack(pady=6)
        ttk.Label(self.tab, textvariable=self.folder_path, font=("微软雅黑", 9)).pack()

        drop_frame = ttk.LabelFrame(self.tab, text="拖拽文件夹到此处", padding="10")
        drop_frame.pack(padx=35, pady=10, fill="x")
        drop_label = ttk.Label(drop_frame, text="点击选择 或 直接拖拽文件夹", font=("微软雅黑", 10))
        drop_label.pack(expand=True)
        drop_label.drop_target_register(DND_FILES)
        drop_label.dnd_bind('<<Drop>>', self.on_drop)

        self.btn_start = ttk.Button(self.tab, text="✅ 生成Excel表格", command=self.generate, 
                                   state=tk.DISABLED, style="Accent.TButton")
        self.btn_start.pack(pady=25, ipadx=30, ipady=6)

    def select_folder(self):
        path = filedialog.askdirectory(title="选择图片文件夹")
        if path:
            self.folder_path.set(path)
            self.btn_start.config(state=tk.NORMAL)

    def on_drop(self, event):
        path = event.data.strip("{}")
        if os.path.isdir(path):
            self.folder_path.set(path)
            self.btn_start.config(state=tk.NORMAL)

    def generate(self):
        folder = self.folder_path.get()
        if not os.path.isdir(folder):
            messagebox.showerror("错误", "请选择有效文件夹")
            return

        title = self.entries["title"].get().strip()
        content = self.entries["content"].get().strip()
        topic = self.entries["topic"].get().strip()
        poster = self.entries["poster"].get().strip()
        instruct = self.entries["instruct"].get().strip()
        
        try:
            per_row = int(self.entries["per_row"].get())
            if not 1 <= per_row <=5:
                messagebox.showerror("错误","请输入1-5之间的数字")
                return
        except:
            messagebox.showerror("错误","每行图片数必须是有效数字")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "图片表格"
            headers = ["序号","标题","文案","话题","大字报","图片1","图片2","图片3",
                       "图片4","图片5","视频","使用说明","指定用户"]
            for i,h in enumerate(headers,1):
                ws.cell(row=1,column=i,value=h)

            exts = ('.jpg','.jpeg','.png','.bmp','.gif')
            imgs = [f for f in os.listdir(folder) if f.lower().endswith(exts)]
            groups = [imgs[i:i+per_row] for i in range(0,len(imgs),per_row)]

            for row_idx, group in enumerate(groups,2):
                ws.cell(row=row_idx,column=1,value=row_idx-1)
                ws.cell(row=row_idx,column=2,value=title)
                ws.cell(row=row_idx,column=3,value=content)
                ws.cell(row=row_idx,column=4,value=topic)
                ws.cell(row=row_idx,column=5,value=poster)
                ws.cell(row=row_idx,column=12,value=instruct)

                for col_idx, img_file in enumerate(group):
                    img_path = os.path.join(folder,img_file)
                    try:
                        img = Image(img_path)
                        w = max(300//per_row,120)
                        img.width = w
                        img.height=200
                        ws.add_image(img,f"{get_column_letter(6+col_idx)}{row_idx}")
                        ws.row_dimensions[row_idx].height=220
                    except:
                        continue

            widths = [6,20,30,15,15,35,35,35,35,35,20,20,15]
            for i,w in enumerate(widths,1):
                ws.column_dimensions[get_column_letter(i)].width=w

            save_path = os.path.join(folder,"图片内容表格.xlsx")
            wb.save(save_path)
            messagebox.showinfo("✅ 成功",f"生成{len(groups)}行，{len(imgs)}张图片")
            os.startfile(folder)
        except Exception as e:
            messagebox.showerror("❌ 失败",str(e))